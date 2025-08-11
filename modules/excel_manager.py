"""
Модуль для работы с Excel файлами - чтение и запись данных.
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime, date
from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def get_date_from_first_row(df: pd.DataFrame) -> date:
    """
    Получает дату из первой строки данных в колонке "ДатаБезВремени".

    Args:
        df: DataFrame с данными

    Returns:
        date: Дата из первой строки данных
    """
    if "ДатаБезВремени" not in df.columns:
        logger.error("❌ Колонка 'ДатаБезВремени' не найдена в данных")
        raise ValueError("Колонка 'ДатаБезВремени' обязательна для автоматического определения даты")

    # Получаем первую непустую дату
    first_date = None
    for idx, row in df.iterrows():
        date_value = row['ДатаБезВремени']
        if pd.notna(date_value) and str(date_value).strip() != '':
            try:
                if isinstance(date_value, str):
                    first_date = datetime.strptime(date_value, "%d.%m.%Y").date()
                elif isinstance(date_value, datetime):
                    first_date = date_value.date()
                else:
                    first_date = pd.to_datetime(date_value).date()
                break
            except Exception as e:
                logger.warning(f"⚠️ Не удалось распарсить дату '{date_value}' в строке {idx}: {e}")
                continue

    if first_date is None:
        logger.error("❌ Не найдена валидная дата в колонке 'ДатаБезВремени'")
        raise ValueError("Не найдена валидная дата в данных")

    logger.info(f"✅ Автоматически определена дата: {first_date.strftime('%d.%m.%Y')}")
    return first_date


def filter_problems_by_date(df: pd.DataFrame, target_date: date) -> pd.DataFrame:
    """
    Фильтрует проблемы по дате из колонки "ДатаБезВремени".

    Args:
        df: DataFrame с проблемами
        target_date: Целевая дата для фильтрации

    Returns:
        pd.DataFrame: Отфильтрованные проблемы
    """
    logger.info(f"🔍 Фильтруем проблемы для даты {target_date.strftime('%d.%m.%Y')}")

    # Преобразуем target_date в строку для сравнения
    target_date_str = target_date.strftime('%d.%m.%Y')

    # ОТЛАДКА: Показываем первые несколько значений дат
    logger.info(f"🔍 Ищем дату: '{target_date_str}'")
    logger.info(f"🔍 Первые 5 значений в колонке 'ДатаБезВремени':")
    for idx, date_val in df['ДатаБезВремени'].head(5).items():
        logger.info(f"   Строка {idx}: '{date_val}' (тип: {type(date_val)})")

    # Улучшенная фильтрация по дате с учетом разных форматов
    def normalize_date(date_val):
        """Нормализует дату к строковому формату DD.MM.YYYY"""
        if pd.isna(date_val):
            return None

        try:
            if isinstance(date_val, str):
                # Если это строка, пробуем распарсить
                return datetime.strptime(date_val, "%d.%m.%Y").strftime("%d.%m.%Y")
            elif isinstance(date_val, datetime):
                # Если это datetime объект
                return date_val.strftime("%d.%m.%Y")
            elif isinstance(date_val, date):
                # Если это date объект
                return date_val.strftime("%d.%m.%Y")
            else:
                # Пробуем pandas to_datetime
                parsed_date = pd.to_datetime(date_val)
                return parsed_date.strftime("%d.%m.%Y")
        except Exception as e:
            logger.warning(f"⚠️ Не удалось нормализовать дату '{date_val}': {e}")
            return str(date_val)

    # Нормализуем даты в DataFrame
    normalized_dates = df['ДатаБезВремени'].apply(normalize_date)

        # Создаем фильтр по дате
    date_filter = normalized_dates == target_date_str

    # Применяем фильтр по дате
    filtered_df = df[date_filter].copy()
    logger.info(f"📊 Найдено {len(filtered_df)} проблем для даты {target_date.strftime('%d.%m.%Y')}")

    if len(filtered_df) == 0:
        logger.warning(f"⚠️ Не найдено проблем для даты {target_date.strftime('%d.%m.%Y')}")
        # Дополнительная отладка
        logger.info(f"🔍 Все уникальные даты в файле:")
        unique_dates = df['ДатаБезВремени'].dropna().unique()
        for date_val in unique_dates[:10]:  # Показываем первые 10
            logger.info(f"   '{date_val}' (тип: {type(date_val)})")
    else:
        logger.info(f"✅ Проблемы для обработки:")
        for idx, row in filtered_df.iterrows():
            date_str = row['ДатаБезВремени'] if pd.notna(row['ДатаБезВремени']) else 'N/A'
            logger.info(f"   📋 {row['Номер массовой']} - {row['Регион']} (дата: {date_str})")

    return filtered_df


def calculate_time_window_for_date(row: pd.Series, target_date: date) -> tuple[datetime, datetime]:
    """
    Вычисляет временное окно для указанной даты с учетом времени открытия/закрытия проблемы.

    Args:
        row: Строка с данными проблемы
        target_date: Целевая дата

    Returns:
        tuple[datetime, datetime]: (время начала, время окончания) для указанной даты
    """
    target_datetime_start = datetime.combine(target_date, datetime.min.time())  # 00:00:00
    target_datetime_end = datetime.combine(target_date, datetime.max.time())    # 23:59:59

    problem_start = row['Старт']
    problem_end = row['Окончание']

    # Определяем время начала для указанной даты
    if problem_start.date() == target_date:
        # Проблема была открыта в указанную дату - берем время открытия
        window_start = problem_start
    else:
        # Проблема была открыта раньше - берем начало дня
        window_start = target_datetime_start

    # Определяем время окончания для указанной даты
    if pd.isna(problem_end):
        # Проблема не закрыта - берем конец дня
        window_end = target_datetime_end
    elif problem_end.date() == target_date:
        # Проблема была закрыта в указанную дату - берем время закрытия
        window_end = problem_end
    else:
        # Проблема была закрыта позже - берем конец дня
        window_end = target_datetime_end

    logger.info(f"🕒 Временное окно для {row['Номер массовой']} на {target_date.strftime('%d.%m.%Y')}:")
    logger.info(f"   📍 Начало: {window_start.strftime('%d.%m.%Y %H:%M')}")
    logger.info(f"   📍 Окончание: {window_end.strftime('%d.%m.%Y %H:%M')}")

    return window_start, window_end


def save_results_to_excel(
    results: List[Dict[str, Any]],
    original_file_path: Path,
    target_date: date
) -> None:
    """
    Сохраняет результаты в отдельный лист Excel файла.
    Создает новый лист с полной таблицей как в исходных данных + колонка "потерянные".

    Args:
        results: Список результатов
        original_file_path: Путь к исходному Excel файлу
        target_date: Дата для которой были получены результаты
    """
    logger.info(f"💾 Сохраняем результаты в отдельный лист: {original_file_path}")

    try:
        # Загружаем рабочую книгу
        workbook = load_workbook(original_file_path)

        # Создаем или получаем лист для результатов
        sheet_name = f"Результаты_{target_date.strftime('%d_%m_%Y')}"
        if sheet_name in workbook.sheetnames:
            result_sheet = workbook[sheet_name]
        else:
            result_sheet = workbook.create_sheet(sheet_name)
            logger.info(f"📋 Создан новый лист: {sheet_name}")

        # Если лист новый, копируем заголовки из исходного листа "Отчет"
        if result_sheet.max_row == 1 and result_sheet.max_column == 1:
            # Копируем заголовки из листа "Отчет"
            report_sheet = workbook["Отчет"]
            for col in range(1, report_sheet.max_column + 1):
                cell_value = report_sheet.cell(row=1, column=col).value
                result_sheet.cell(row=1, column=col, value=cell_value)

            # Проверяем есть ли уже колонки "Потерянные" и "Превышение"
            lost_col = None
            excess_col = None

            for col in range(1, result_sheet.max_column + 1):
                header = result_sheet.cell(row=1, column=col).value
                if header:
                    header_str = str(header).strip()
                    if "потерянн" in header_str.lower():
                        lost_col = col
                    elif "превышен" in header_str.lower():
                        excess_col = col

            # Добавляем колонки только если их нет
            if lost_col is None:
                lost_col = result_sheet.max_column + 1
                result_sheet.cell(row=1, column=lost_col, value="Потерянные")
                logger.info(f"📋 Добавлена колонка 'Потерянные' в позицию {lost_col}")
            else:
                logger.info(f"📋 Найдена существующая колонка 'Потерянные' в позиции {lost_col}")

            if excess_col is None:
                excess_col = result_sheet.max_column + 1
                result_sheet.cell(row=1, column=excess_col, value="Превышение")
                logger.info(f"📋 Добавлена колонка 'Превышение' в позицию {excess_col}")
            else:
                logger.info(f"📋 Найдена существующая колонка 'Превышение' в позиции {excess_col}")

            logger.info(f"📋 Скопированы заголовки и проверены колонки для результатов")

        # Находим колонку с номером массовой
        mass_number_col = None
        for col in range(1, result_sheet.max_column + 1):
            header = result_sheet.cell(row=1, column=col).value
            if header and "номер" in str(header).lower() and "массовой" in str(header).lower():
                mass_number_col = col
                break

        if mass_number_col is None:
            logger.error("❌ Не найдена колонка с номером массовой")
            return

        # Создаем словарь результатов
        results_dict = {}
        for result in results:
            mass_number = result["Номер массовой"]
            lost_calls = result["LostCalls"]
            excess_traffic = result.get("ExcessTraffic", 0.0)  # Добавляем поддержку excess_traffic
            results_dict[mass_number] = {"lost": lost_calls, "excess": excess_traffic}

        # Обрабатываем каждый результат
        for mass_number, data in results_dict.items():
            lost_calls = data["lost"]
            excess_traffic = data["excess"]

            # Ищем строку с нужным номером массовой
            target_row = None
            for row in range(2, result_sheet.max_row + 1):
                cell_value = result_sheet.cell(row=row, column=mass_number_col).value
                if str(cell_value) == str(mass_number):
                    target_row = row
                    break

            if target_row is None:
                # Добавляем новую строку
                target_row = result_sheet.max_row + 1
                logger.info(f"➕ Добавлена новая строка {target_row} для {mass_number}")

            # Копируем данные из исходного листа "Отчет"
            report_sheet = workbook["Отчет"]
            report_row = None
            for row in range(2, report_sheet.max_row + 1):
                cell_value = report_sheet.cell(row=row, column=mass_number_col).value
                if str(cell_value) == str(mass_number):
                    report_row = row
                    break

            if report_row:
                # Копируем все данные из исходной строки
                for col in range(1, report_sheet.max_column + 1):
                    cell_value = report_sheet.cell(row=report_row, column=col).value
                    result_sheet.cell(row=target_row, column=col, value=cell_value)

            # Записываем результаты в найденные колонки
            if lost_col is not None:
                result_sheet.cell(row=target_row, column=lost_col, value=lost_calls)
            if excess_col is not None:
                result_sheet.cell(row=target_row, column=excess_col, value=excess_traffic)

            logger.info(f"✅ Сохранен результат: {mass_number} → lost={lost_calls}, excess={excess_traffic}")

        # Сохраняем файл
        workbook.save(original_file_path)
        logger.info(f"💾 Все результаты сохранены в лист '{sheet_name}'")

    except Exception as e:
        logger.error(f"❌ Ошибка при сохранении результатов: {e}")


def check_notes_column(
    workbook,
    report_sheet,
    target_row: int
) -> tuple:
    """
    Проверяет значение в колонке "Заметки" для определения необходимости обработки.

    Args:
        workbook: Рабочая книга Excel
        report_sheet: Лист для проверки
        target_row: Номер строки для проверки

    Returns:
        tuple: (should_skip, notes_value, reason) - нужно ли пропускать строку
    """
    try:
        # Ищем колонку "Заметки"
        notes_col = None
        for col in range(1, report_sheet.max_column + 1):
            header = report_sheet.cell(row=1, column=col).value
            if header and "заметк" in str(header).lower():
                notes_col = col
                break

        if notes_col is None:
            logger.warning("⚠️ Колонка 'Заметки' не найдена - обрабатываем все строки")
            return False, None, "notes_column_not_found"

        # Получаем значение из колонки "Заметки"
        notes_value = report_sheet.cell(row=target_row, column=notes_col).value

        if notes_value is None:
            logger.info(f"📝 Строка {target_row}: колонка 'Заметки' пустая - обрабатываем")
            return False, notes_value, "notes_empty"

        # Пытаемся преобразовать в число для сравнения
        try:
            notes_numeric = float(str(notes_value).strip())
            if notes_numeric < 50:
                logger.info(f"📝 Строка {target_row}: заметка = {notes_value} < 50 - пропускаем вычисления")
                return True, notes_value, "notes_less_50"
            else:
                logger.info(f"📝 Строка {target_row}: заметка = {notes_value} >= 50 - обрабатываем")
                return False, notes_value, "notes_ok"
        except (ValueError, TypeError):
            # Если не удалось преобразовать в число, считаем что это текстовая заметка
            logger.info(f"📝 Строка {target_row}: заметка = '{notes_value}' (текст) - обрабатываем")
            return False, notes_value, "notes_text"

    except Exception as e:
        logger.error(f"❌ Ошибка при проверке колонки 'Заметки': {e}")
        # В случае ошибки обрабатываем строку
        return False, None, "error_checking_notes"


def save_excel_batch(
    workbook,
    report_sheet,
    original_file_path: Path,
    batch_size: int = 10
) -> bool:
    """
    Сохраняет Excel файл пакетно.

    Args:
        workbook: Рабочая книга Excel
        report_sheet: Лист для сохранения
        original_file_path: Путь к файлу для сохранения
        batch_size: Размер пакета для сохранения

    Returns:
        bool: True если сохранение успешно
    """
    try:
        workbook.save(original_file_path)
        logger.info(f"💾 Пакет сохранен в файл: {original_file_path.name}")
        return True
    except PermissionError as pe:
        logger.error(f"❌ ОШИБКА ДОСТУПА: Файл {original_file_path} заблокирован")
        logger.error(f"   Возможные причины:")
        logger.error(f"   - Файл открыт в Excel")
        logger.error(f"   - Файл открыт в другой программе")
        logger.error(f"   - Недостаточно прав доступа")
        raise pe
    except Exception as save_e:
        logger.error(f"❌ Ошибка при сохранении файла: {save_e}")
        raise save_e


def save_single_result_to_original_file(
    mass_number: str,
    lost_calls: int,
    excess_traffic: float,
    original_file_path: Path,
    row_index: int,
    workbook=None,
    report_sheet=None,
    save_counter=0
) -> tuple:
    """
    Сохраняет результат одной строки в исходный Excel файл.
    Добавляет колонки "Потерянные" и "Превышение" если их нет.

    Args:
        mass_number: Номер массового инцидента
        lost_calls: Количество потерянных звонков
        excess_traffic: Коэффициент превышения трафика (сохраняется как "Превышение")
        original_file_path: Путь к исходному Excel файлу
        row_index: Индекс строки в исходном файле
        workbook: Опционально - уже открытая рабочая книга (для пакетного сохранения)
        report_sheet: Опционально - уже открытый лист (для пакетного сохранения)
        save_counter: Счетчик записей для пакетного сохранения

    Returns:
        tuple: (workbook, report_sheet, save_counter, skip_reason) для возможного повторного использования
    """
    logger.info(f"💾 Сохраняем результат для {mass_number}: lost={lost_calls}, excess={excess_traffic}")

    try:
        # Загружаем рабочую книгу, если не передана
        if workbook is None:
            workbook = load_workbook(original_file_path)
            report_sheet = workbook["Отчет"]
        elif report_sheet is None:
            report_sheet = workbook["Отчет"]

        # Проверяем есть ли колонки "Потерянные" и "Превышение"
        lost_col = None
        excess_col = None

        for col in range(1, report_sheet.max_column + 1):
            header = report_sheet.cell(row=1, column=col).value
            if header:
                header_str = str(header).strip()
                if "потерянн" in header_str.lower():
                    lost_col = col
                    logger.info(f"🔍 Найдена существующая колонка 'Потерянные' в позиции {col}")
                elif "превышен" in header_str.lower():
                    excess_col = col
                    logger.info(f"🔍 Найдена существующая колонка 'Превышение' в позиции {col}")

        # Если колонок нет, добавляем их
        if lost_col is None:
            lost_col = report_sheet.max_column + 1
            report_sheet.cell(row=1, column=lost_col, value="Потерянные")
            logger.info(f"➕ Добавлена колонка 'Потерянные' в позицию {lost_col}")

        if excess_col is None:
            excess_col = report_sheet.max_column + 1
            report_sheet.cell(row=1, column=excess_col, value="Превышение")
            logger.info(f"➕ Добавлена колонка 'Превышение' в позицию {excess_col}")

        # Находим строку с нужным номером массовой
        mass_number_col = None
        for col in range(1, report_sheet.max_column + 1):
            header = report_sheet.cell(row=1, column=col).value
            if header and "номер" in str(header).lower() and "массовой" in str(header).lower():
                mass_number_col = col
                break

        if mass_number_col is None:
            logger.error("❌ Не найдена колонка с номером массовой")
            return workbook, report_sheet, save_counter, "no_mass_number_col"

        # Ищем строку с нужным номером массовой
        target_row = None
        for row in range(2, report_sheet.max_row + 1):
            cell_value = report_sheet.cell(row=row, column=mass_number_col).value
            if str(cell_value) == str(mass_number):
                target_row = row
                break

        if target_row is None:
            logger.error(f"❌ Не найдена строка с номером массовой {mass_number}")
            return workbook, report_sheet, save_counter, "row_not_found"

        # Проверяем, есть ли уже значения в колонках "Потерянные" и "Превышение"
        existing_lost = report_sheet.cell(row=target_row, column=lost_col).value
        existing_excess = report_sheet.cell(row=target_row, column=excess_col).value

        if existing_lost is not None and existing_excess is not None:
            # Проверяем, что значения не пустые и не равны 0 (если это не специальный случай)
            if str(existing_lost).strip() != "" and str(existing_excess).strip() != "":
                logger.info(f"⏭️ Строка {mass_number} уже обработана (lost={existing_lost}, excess={existing_excess}) - пропускаем")
                return workbook, report_sheet, save_counter, "already_processed"

        # Записываем результаты в существующие колонки
        report_sheet.cell(row=target_row, column=lost_col, value=lost_calls)
        report_sheet.cell(row=target_row, column=excess_col, value=excess_traffic)

        # Увеличиваем счетчик записей
        save_counter += 1

        logger.info(f"✅ Результат записан в строку {target_row}: {mass_number} → lost={lost_calls} (колонка {lost_col}), excess={excess_traffic} (колонка {excess_col})")

        return workbook, report_sheet, save_counter, "success"

    except Exception as e:
        logger.error(f"❌ Ошибка при записи результата для {mass_number}: {e}")
        raise