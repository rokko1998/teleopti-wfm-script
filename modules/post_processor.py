"""
Модуль для постобработки данных после выгрузки.

Выполняет:
1. Поиск дубликатов в колонке "Потерянные"
2. Сравнение регионов и заметок для одинаковых значений
3. Оставление только строки с наибольшим количеством заметок для каждого региона
4. Зануление "Потерянные" для строк с отрицательным "Превышение"
5. Логирование всего процесса
"""

from pathlib import Path
from typing import Dict, List, Tuple, Any
import pandas as pd
from openpyxl import load_workbook
from loguru import logger


def post_process_excel_file(file_path: Path) -> None:
    """
    Выполняет постобработку Excel файла после выгрузки данных.

    Args:
        file_path: Путь к Excel файлу для обработки
    """
    logger.info(f"🔧 Начинаем постобработку файла: {file_path}")

    try:
        # Загружаем рабочую книгу
        workbook = load_workbook(file_path)
        report_sheet = workbook["Отчет"]

        # Находим колонки
        column_mapping = _find_columns(report_sheet)
        logger.info(f"📋 Найдены колонки: {column_mapping}")

        # Читаем данные в DataFrame для удобной обработки
        df = _read_sheet_to_dataframe(report_sheet, column_mapping)
        logger.info(f"📊 Загружено {len(df)} строк данных")

        # Выполняем постобработку
        df_processed = _process_duplicates(df, column_mapping)
        df_processed = _process_negative_excess(df_processed, column_mapping)

        # Сохраняем изменения обратно в Excel
        _save_dataframe_to_sheet(df_processed, report_sheet, column_mapping)

        # Сохраняем файл
        workbook.save(file_path)
        logger.info(f"✅ Постобработка завершена успешно. Файл сохранен: {file_path}")

    except Exception as e:
        logger.error(f"❌ Ошибка при постобработке файла {file_path}: {e}")
        logger.exception("Полный traceback:")
        raise


def _find_columns(sheet) -> Dict[str, int]:
    """
    Находит индексы нужных колонок в листе.

    Args:
        sheet: Лист Excel файла

    Returns:
        Dict с названиями колонок и их индексами
    """
    column_mapping = {}

    # Ищем колонки по заголовкам
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header:
            header_str = str(header).strip().lower()

            if "потерянн" in header_str:
                column_mapping["Потерянные"] = col
            elif "регион" in header_str:
                column_mapping["Регион"] = col
            elif "заметк" in header_str:
                column_mapping["Заметки"] = col
            elif "превышен" in header_str:
                column_mapping["Превышение"] = col
            elif "номер массовой" in header_str:
                column_mapping["Номер массовой"] = col

    # Проверяем что все нужные колонки найдены
    required_columns = ["Потерянные", "Регион", "Превышение"]
    missing_columns = [col for col in required_columns if col not in column_mapping]

    if missing_columns:
        raise ValueError(f"Не найдены обязательные колонки: {missing_columns}")

    # Заметки могут отсутствовать - это нормально
    if "Заметки" not in column_mapping:
        logger.warning("⚠️ Колонка 'Заметки' не найдена, будет использоваться пустая строка")
        column_mapping["Заметки"] = None

    return column_mapping


def _read_sheet_to_dataframe(sheet, column_mapping: Dict[str, int]) -> pd.DataFrame:
    """
    Читает данные из листа в DataFrame.

    Args:
        sheet: Лист Excel файла
        column_mapping: Словарь с колонками и их индексами

    Returns:
        DataFrame с данными
    """
    data = []

    # Читаем данные начиная со второй строки (первая - заголовки)
    for row in range(2, sheet.max_row + 1):
        row_data = {}

        for col_name, col_idx in column_mapping.items():
            if col_idx is not None:
                cell_value = sheet.cell(row=row, column=col_idx).value
                row_data[col_name] = cell_value
            else:
                row_data[col_name] = ""  # Для отсутствующих колонок

        # Добавляем индекс строки для отслеживания
        row_data["_row_index"] = row
        data.append(row_data)

    return pd.DataFrame(data)


def _process_duplicates(df: pd.DataFrame, column_mapping: Dict[str, int]) -> pd.DataFrame:
    """
    Обрабатывает дубликаты в колонке "Потерянные".

    Args:
        df: DataFrame с данными
        column_mapping: Словарь с колонками и их индексами

    Returns:
        Обработанный DataFrame
    """
    logger.info("🔍 Начинаем поиск дубликатов в колонке 'Потерянные'")

    # Группируем по значениям "Потерянные"
    lost_groups = df.groupby("Потерянные")

    total_duplicates = 0
    processed_groups = 0

    for lost_value, group in lost_groups:
        if pd.isna(lost_value) or lost_value == 0:
            continue  # Пропускаем пустые и нулевые значения

        if len(group) == 1:
            continue  # Нет дубликатов

        logger.info(f"🔍 Найдена группа дубликатов для значения '{lost_value}': {len(group)} строк")

        # Группируем по регионам внутри группы дубликатов
        region_groups = group.groupby("Регион")

        for region, region_group in region_groups:
            if len(region_group) == 1:
                continue  # Нет дубликатов в регионе

            logger.info(f"   📍 Регион '{region}': {len(region_group)} дубликатов")

            # Сравниваем заметки и оставляем строку с наибольшим количеством
            best_row_idx = _find_best_row_by_notes(region_group)

            # Зануляем "Потерянные" для всех остальных строк в этой группе
            for idx, row in region_group.iterrows():
                if idx != best_row_idx:
                    df.loc[idx, "Потерянные"] = 0
                    total_duplicates += 1
                    logger.info(f"      🗑️ Занулена строка {row['_row_index']} (массовая: {row.get('Номер массовой', 'N/A')})")
                else:
                    logger.info(f"      ✅ Оставлена строка {row['_row_index']} (массовая: {row.get('Номер массовой', 'N/A')}) - наибольшее количество заметок")

            processed_groups += 1

    logger.info(f"✅ Обработка дубликатов завершена: {processed_groups} групп, {total_duplicates} строк занулено")
    return df


def _find_best_row_by_notes(group: pd.DataFrame) -> int:
    """
    Находит строку с наибольшим количеством заметок в группе.
    
    Логика выбора:
    1. Если есть колонка "Заметки" - выбираем строку с наибольшим количеством заметок
    2. Если колонки "Заметки" нет - выбираем строку с наибольшим "Номер массовой"
    3. Если и "Номер массовой" одинаковые - берем первую строку
    
    Args:
        group: Группа строк с одинаковым регионом и значением "Потерянные"
        
    Returns:
        Индекс строки с наибольшим количеством заметок
    """
    if "Заметки" not in group.columns:
        # Если колонки "Заметки" нет, используем "Номер массовой" как критерий
        logger.info("   📝 Колонка 'Заметки' отсутствует, используем 'Номер массовой' для выбора")
        return _find_best_row_by_mass_number(group)
    
    # Подсчитываем количество заметок для каждой строки
    notes_counts = []
    for idx, row in group.iterrows():
        notes = row["Заметки"]
        if pd.isna(notes) or notes == "":
            count = 0
        else:
            # Улучшенный подсчет заметок:
            # 1. Количество строк (разделенных \n)
            # 2. Количество слов (разделенных пробелами)
            # 3. Количество символов
            notes_str = str(notes).strip()
            line_count = len(notes_str.split('\n')) if '\n' in notes_str else 1
            word_count = len(notes_str.split()) if notes_str else 0
            char_count = len(notes_str)
            
            # Комбинированная оценка: приоритет количеству строк, затем слов, затем символов
            count = line_count * 1000 + word_count * 10 + char_count
            
            logger.info(f"      📝 Строка {row['_row_index']}: {line_count} строк, {word_count} слов, {char_count} символов (оценка: {count})")
        
        notes_counts.append((count, idx))
    
    # Сортируем по убыванию количества заметок
    notes_counts.sort(key=lambda x: x[0], reverse=True)
    
    best_idx = notes_counts[0][1]
    best_row = group.loc[best_idx]
    logger.info(f"   ✅ Выбрана строка {best_row['_row_index']} с наибольшим количеством заметок")
    
    return best_idx


def _find_best_row_by_mass_number(group: pd.DataFrame) -> int:
    """
    Находит строку с наибольшим номером массовой в группе.
    
    Args:
        group: Группа строк с одинаковым регионом и значением "Потерянные"
        
    Returns:
        Индекс строки с наибольшим номером массовой
    """
    if "Номер массовой" not in group.columns:
        # Если и колонки "Номер массовой" нет, берем первую строку
        logger.info("   📝 Колонка 'Номер массовой' отсутствует, выбираем первую строку")
        return group.index[0]
    
    # Ищем строку с наибольшим номером массовой
    max_mass_number = None
    best_idx = group.index[0]
    
    for idx, row in group.iterrows():
        mass_number = row["Номер массовой"]
        if pd.isna(mass_number):
            continue
            
        # Пытаемся извлечь числовую часть из номера массовой
        try:
            # Если номер содержит числа, извлекаем их
            import re
            numbers = re.findall(r'\d+', str(mass_number))
            if numbers:
                numeric_value = int(''.join(numbers))
                if max_mass_number is None or numeric_value > max_mass_number:
                    max_mass_number = numeric_value
                    best_idx = idx
                    logger.info(f"      📝 Строка {row['_row_index']}: номер массовой '{mass_number}' -> {numeric_value}")
        except (ValueError, TypeError):
            # Если не удалось извлечь число, сравниваем как строки
            if max_mass_number is None or str(mass_number) > str(max_mass_number):
                max_mass_number = mass_number
                best_idx = idx
                logger.info(f"      📝 Строка {row['_row_index']}: номер массовой '{mass_number}' (строковое сравнение)")
    
    logger.info(f"   ✅ Выбрана строка {group.loc[best_idx]['_row_index']} с наибольшим номером массовой")
    return best_idx


def _process_negative_excess(df: pd.DataFrame, column_mapping: Dict[str, int]) -> pd.DataFrame:
    """
    Зануляет "Потерянные" для строк с отрицательным "Превышение".

    Args:
        df: DataFrame с данными
        column_mapping: Словарь с колонками и их индексами

    Returns:
        Обработанный DataFrame
    """
    logger.info("🔍 Начинаем обработку строк с отрицательным 'Превышение'")

    # Находим строки с отрицательным превышением
    negative_excess_mask = pd.to_numeric(df["Превышение"], errors='coerce') < 0

    negative_count = negative_excess_mask.sum()

    if negative_count == 0:
        logger.info("✅ Строк с отрицательным 'Превышение' не найдено")
        return df

    logger.info(f"🔍 Найдено {negative_count} строк с отрицательным 'Превышение'")

    # Зануляем "Потерянные" для этих строк
    df.loc[negative_excess_mask, "Потерянные"] = 0

    # Логируем детали
    negative_rows = df[negative_excess_mask]
    for idx, row in negative_rows.iterrows():
        logger.info(f"   🗑️ Занулена строка {row['_row_index']} (массовая: {row.get('Номер массовой', 'N/A')}, превышение: {row['Превышение']})")

    logger.info(f"✅ Обработка отрицательных превышений завершена: {negative_count} строк занулено")
    return df


def _save_dataframe_to_sheet(df: pd.DataFrame, sheet, column_mapping: Dict[str, int]) -> None:
    """
    Сохраняет обработанный DataFrame обратно в лист Excel.

    Args:
        df: Обработанный DataFrame
        sheet: Лист Excel файла
        column_mapping: Словарь с колонками и их индексами
    """
    logger.info("💾 Сохраняем обработанные данные в Excel файл")

    for idx, row in df.iterrows():
        row_index = row["_row_index"]

        # Обновляем только измененные колонки
        for col_name, col_idx in column_mapping.items():
            if col_idx is not None and col_name in df.columns:
                sheet.cell(row=row_index, column=col_idx, value=row[col_name])

    logger.info("✅ Данные успешно сохранены в Excel файл")
